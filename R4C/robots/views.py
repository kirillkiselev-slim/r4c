import json
from .models import Robot
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from .serializers import RobotSerializer
import openpyxl
from django.db.models import Count
from datetime import timedelta
from django.utils import timezone


@csrf_exempt
def create_robot(request):
    if request.method == "POST":
        dt = json.loads(request.body)
        serial = dt.get('serial')
        model = dt.get('model')
        version = dt.get('version')
        created = dt.get('created')

        serializer = RobotSerializer(data=dt)

        if serializer.is_valid():

            if serial and model and version and created:
                robot = Robot(serial=serial.upper(), model=model.upper(),
                              version=version.upper(), created=created)
                robot.save()
                return JsonResponse({"success": True})
            else:
                return JsonResponse({"success": False, "error": "Invalid data"})
        else:
            return JsonResponse({"success": False, "error": serializer.errors})
    else:
        return JsonResponse({"success": False, "error": "Invalid request method"})


def generate_excel_report():
    filepath = "../robots_reports.xlsx"
    last_week = timezone.now() - timedelta(days=7)

    wb = openpyxl.Workbook()
    default_sheet = wb.active
    excel_column_names = ['Модель', 'Версия', 'Количество за неделю']

    for col_num, col_name in enumerate(excel_column_names, start=1):
        default_sheet.cell(row=1,column=col_num).value = col_name

    data = Robot.objects.filter(created__gte=last_week).values_list('model', 'version').annotate(count=Count('version'))

    models = set(row[0] for row in data)
    for model in models:
        new_sheet = wb.create_sheet(title=model)
        new_sheet.append(excel_column_names)

        for row_data in data:
            if row_data[0] == model:
                new_sheet.append(row_data)

    wb.remove(wb.active)

    wb.save(filepath)

    return filepath

report = generate_excel_report()


@csrf_exempt
def download_excel_report(request):
    file_path = "../robots_reports.xlsx"
    with open(file_path, 'rb') as file:
        response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=robots_reports.xlsx'
    return response

